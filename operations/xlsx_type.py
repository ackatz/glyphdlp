from openpyxl import load_workbook
from io import BytesIO
from app.detections import regex
import re
import base64


def scan(input) -> dict:
    """
    Decode Base64 data and scan the XLSX content.
    :param input: Base64-encoded XLSX data
    :return: The results of the scan
    """

    wb = load_workbook(BytesIO(input), read_only=True)

    results = {}

    for sheet in wb:
        for row in sheet.iter_rows(values_only=True):
            for cell_value in row:
                if cell_value is not None and isinstance(cell_value, str):
                    for name, pattern in regex.patterns.items():
                        matches = re.findall(pattern, cell_value)
                        if matches:
                            if name not in results:
                                results[name] = []
                            results[name].extend(matches)

    return results


def redact(input) -> str:
    """
    Redact sensitive information in Base64-encoded XLSX data and encode back to Base64.
    :param input: Base64-encoded XLSX data
    :return: Base64-encoded redacted XLSX data
    """

    wb = load_workbook(BytesIO(input))

    for sheet in wb:
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is not None and isinstance(cell.value, str):
                    for name, pattern in regex.patterns.items():
                        if re.search(pattern, cell.value):
                            cell.value = re.sub(pattern, "[REDACTED]", cell.value)

    output_buffer = BytesIO()
    wb.save(output_buffer)
    redacted_input = base64.b64encode(output_buffer.getvalue()).decode("utf-8")

    return redacted_input
